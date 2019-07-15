"""
adder.py
~~~~~~

Creates a simple GUI for summing two numbers.
"""

import tkinter
from tkinter import ttk, BooleanVar, Frame, Tk, BOTH, Text, Menu, END
from CalcLib import CalcLib
import openpyxl
from openpyxl import load_workbook
import tkinter.filedialog
import numpy as np
import matplotlib.pyplot as plt


class Adder(ttk.Frame):
    """The adders gui and functions."""
    G1 = "English"
    G2 = "Spahish"
    G3 = "Photography"
    G4 = "Art"
    G5 = "Broadcast"
    G6 = "Calculus"
    G7 = "NASA"


    def __init__(self, parent, *args, **kwargs):
        ttk.Frame.__init__(self, parent, *args, **kwargs)
        self.root = parent
        self.init_gui()

    def on_quit(self):
        """Exits program."""
        quit()

    def calculate(self):
        """Calculates the sum of the two inputted numbers."""
        numbers = [0, 0, 0, 0, 0, 0, 0]
        numbers[0] = int(self.num1_entry.get())
        numbers[1] = int(self.num2_entry.get())
        numbers[2] = int(self.num3_entry.get())
        numbers[3] = int(self.num4_entry.get())
        numbers[4] = int(self.num5_entry.get())
        numbers[5] = int(self.num6_entry.get())
        numbers[6] = int(self.num7_entry.get())

        myCalcLib = CalcLib()
        print(numbers)
        myCalcLib.grade_calc(numbers)
        myCalcLib.advisory_calc(self.var.get())
        total = myCalcLib.get_grades()
        self.answer_label['text'] = total
        return total

    def export(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = "English"
        sheet['B1'] = self.num1_entry.get()
        sheet['A2'] = "Spanish"
        sheet['B2'] = self.num2_entry.get()
        sheet['A3'] = "Physics"
        sheet['B3'] = self.num3_entry.get()
        sheet['A4'] = "History"
        sheet['B4'] = self.num4_entry.get()
        sheet['A5'] = "Creative Writing"
        sheet['B5'] = self.num5_entry.get()
        sheet['A6'] = "Trigenometry"
        sheet['B6'] = self.num6_entry.get()
        sheet['A7'] = "Chemistry"
        sheet['B7'] = self.num7_entry.get()
        sheet['A8'] = "Total"
        sheet['B8'] = self.calculate()
        wb.save('Grades.xlsx')

    def import_file(self):
        wb = load_workbook('Grades.xlsx')
        sheet = wb.active
        self.num1_entry.insert('0', sheet['B1'].value)
        self.num2_entry.insert('0', sheet['B2'].value)
        self.num3_entry.insert('0', sheet['B3'].value)
        self.num4_entry.insert('0', sheet['B4'].value)
        self.num5_entry.insert('0', sheet['B5'].value)
        self.num6_entry.insert('0', sheet['B6'].value)
        self.num7_entry.insert('0', sheet['B7'].value)
        # self.num1_entry = ttk.Entry(self, width=5, textvariable = numbers[0])

    def save_as(self):
        dlg = tkinter.filedialog.SaveAs(self)
        filename = dlg.show()
        f = open(filename, "w")
        f.write(self.num1_entry.get()+ '\n')
        f.write(self.num2_entry.get()+ '\n')
        f.write(self.num3_entry.get()+ '\n')
        f.write(self.num4_entry.get()+ '\n')
        f.write(self.num5_entry.get()+ '\n')
        f.write(self.num6_entry.get()+ '\n')
        f.write(self.num7_entry.get()+ '\n')

    def load_file(self):
        dlg = tkinter.filedialog.Open(self)
        filename = dlg.show()
        f = open(filename, "r")
        self.num1_entry.insert('0',f.readline())
        self.num2_entry.insert('0',f.readline())
        self.num3_entry.insert('0',f.readline())
        self.num4_entry.insert('0',f.readline())
        self.num5_entry.insert('0',f.readline())
        self.num6_entry.insert('0',f.readline())
        self.num7_entry.insert('0',f.readline())

    def graph(self):
        N = 7
        grades = (int(self.num1_entry.get()),
                  int(self.num2_entry.get()),
                  int(self.num3_entry.get()),
                  int(self.num4_entry.get()),
                  int(self.num5_entry.get()),
                  int(self.num6_entry.get()),
                  int(self.num7_entry.get()))

        ind = np.arange(N)  # the x locations for the groups
        width = 0.35       # the width of the bars

        fig, ax = plt.subplots()
        rects1 = ax.bar(ind, grades, width, color='b', yerr=0)

        # add some text for labels, title and axes ticks
        ax.set_ylabel('Grades')
        ax.set_title('Yes')
        ax.set_xticks(ind + width / 2)
        ax.set_xticklabels((self.G1, self.G2, self.G3, self.G4, self.G5, self.G6, self.G7))

        self.autolabel(rects1, ax)
        plt.show()

    def autolabel(self, rects, ax):
        """
        Attach a text label above each bar displaying its height
        """
        for rect in rects:
            height = rect.get_height()
            ax.text(rect.get_x() + rect.get_width()/2., 1.05*height,
                    '%d' % int(height),
                    ha='center', va='bottom')


    def init_gui(self):
        """Builds GUI."""
        self.root.title('Number Adder')
        self.root.option_add('*tearOff', 'FALSE')

        self.grid(column=0, row=0, sticky='nsew')

        self.menubar = tkinter.Menu(self.root)

        self.menu_file = tkinter.Menu(self.menubar)
        self.menu_file.add_command(label='Exit', command=self.on_quit)

        self.menu_edit = tkinter.Menu(self.menubar)

        self.menubar.add_cascade(menu=self.menu_file, label='File')
        self.menubar.add_cascade(menu=self.menu_edit, label='Edit')

        self.root.config(menu=self.menubar)

        self.num1_entry = ttk.Entry(self, width=5)
        self.num1_entry.grid(column=1, row=2)

        self.num2_entry = ttk.Entry(self, width=5)
        self.num2_entry.grid(column=1, row=3)

        self.num3_entry = ttk.Entry(self, width=5)
        self.num3_entry.grid(column=1, row=4)

        self.num4_entry = ttk.Entry(self, width=5)
        self.num4_entry.grid(column=1, row=5)

        self.num5_entry = ttk.Entry(self, width=5)
        self.num5_entry.grid(column=1, row=6)

        self.num6_entry = ttk.Entry(self, width=5)
        self.num6_entry.grid(column=1, row=7)

        self.num7_entry = ttk.Entry(self, width=5)
        self.num7_entry.grid(column=1, row=8)

        self.var = tkinter.BooleanVar(False)
        self.adv_ck_entry = ttk.Checkbutton(self, width=5, variable = self.var, onvalue=True, offvalue=False)

        self.adv_ck_entry.grid(column=1, row=9)

        self.calc_button = ttk.Button(self, text='Calculate',
                                      command=self.calculate)
        self.calc_button.grid(column=0, row=10, columnspan=4)

        self.answer_frame = ttk.LabelFrame(self, text='Answer',
                                           height=100)
        self.answer_frame.grid(column=0, row=11, columnspan=4, sticky='nesw')

        self.answer_label = ttk.Label(self.answer_frame, text='')
        self.answer_label.grid(column=0, row=0)

        self.Export_button = ttk.Button(self, text='Export',
                                        command=self.export)
        self.Export_button.grid(column=0, row=12, columnspan=4)

        self.Import_button = ttk.Button(self, text='Import',
                                        command=self.import_file)
        self.Import_button.grid(column=2, row=12, columnspan=4)

        self.Save_As = ttk.Button(self, text='Save As',
                                  command=self.save_as)
        self.Save_As.grid(column=0, row=13, columnspan=4)

        self.Load_button = ttk.Button(self, text='Load',
                                      command=self.load_file)
        self.Load_button.grid(column=2, row=13, columnspan=4)

        self.Graph_button = ttk.Button(self, text='Graph',
                                      command=self.graph)
        self.Graph_button.grid(column=2, row=14, columnspan=4)

        # Labels that remain constant throughout execution.
        ttk.Label(self, text='Number Adder').grid(column=0, row=0,
                                                  columnspan=4)
        ttk.Label(self, text=self.G1).grid(column=0, row=2,
                                             sticky='w')
        ttk.Label(self, text=self.G2).grid(column=0, row=3,
                                             sticky='w')
        ttk.Label(self, text=self.G3).grid(column=0, row=4,
                                          sticky='w')
        ttk.Label(self, text=self.G4).grid(column=0, row=5,
                                             sticky='w')
        ttk.Label(self, text=self.G5).grid(column=0, row=6,
                                            sticky='w')
        ttk.Label(self, text=self.G6).grid(column=0, row=7,
                                                   sticky='w')
        ttk.Label(self, text=self.G7).grid(column=0, row=8,
                                                sticky='w')
        ttk.Label(self, text='Advisory').grid(column=0, row=9,
                                              sticky='w')

        ttk.Separator(self, orient='horizontal').grid(column=0,
                                                      row=1, columnspan=4, sticky='ew')

        for child in self.winfo_children():
            child.grid_configure(padx=5, pady=5)


if __name__ == '__main__':
    root = tkinter.Tk()
    Adder(root)
    root.mainloop()
